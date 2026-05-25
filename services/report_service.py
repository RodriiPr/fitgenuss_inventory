import io
from datetime import datetime
from typing import List
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

from models.ingredient import Ingredient
from models.movement import InventoryMovement
from repositories.ingredient_repository import IngredientRepository
from repositories.movement_repository import MovementRepository
from config.settings import CURRENCY

class FitGenussPDF(FPDF):
    """
    Generador de PDF personalizado con cabecera y pie de página profesionales 
    que siguen la identidad visual de FitGenuss.
    """
    def __init__(self, title_report: str):
        super().__init__()
        self.title_report = title_report

    def header(self):
        # Fondo decorativo superior
        self.set_fill_color(124, 179, 66)  # Verde Matcha
        self.rect(0, 0, 210, 15, 'F')
        
        # Título principal en blanco sobre el verde
        self.set_y(3)
        self.set_font("helvetica", "B", 14)
        self.set_text_color(255, 255, 255)
        self.cell(0, 10, "FITGENUSS - REPOSTERÍA SALUDABLE", border=False, ln=True, align="C")
        
        # Subtítulo y fecha de generación
        self.set_y(18)
        self.set_font("helvetica", "B", 12)
        self.set_text_color(44, 62, 80) # Gris oscuro
        self.cell(120, 8, self.title_report, border=False, ln=False, align="L")
        
        self.set_font("helvetica", "", 9)
        self.set_text_color(127, 140, 141)
        fecha_str = datetime.now().strftime("%d/%m/%Y %H:%M")
        self.cell(0, 8, f"Fecha: {fecha_str}", border=False, ln=True, align="R")
        
        # Línea de separación
        self.set_draw_color(200, 200, 200)
        self.line(10, 27, 200, 27)
        self.ln(5)

    def footer(self):
        # Pie de página
        self.set_y(-15)
        self.set_font("helvetica", "I", 8)
        self.set_text_color(127, 140, 141)
        # Línea decorativa inferior
        self.line(10, 282, 200, 282)
        self.cell(0, 10, f"Página {self.page_no()}/{{nb}} - FitGenuss Inventory System v1.0", border=False, align="C")


class ReportService:
    """
    Servicio de generación de reportes y exportación en formatos PDF y Excel.
    """
    def __init__(self):
        self.ingredient_repo = IngredientRepository()
        self.movement_repo = MovementRepository()

    # --- REPORTES EN EXCEL (OpenPyXL) ---
    def generate_inventory_excel(self, db: Session) -> bytes:
        """
        Genera un reporte del inventario actual en formato Excel (.xlsx).
        """
        ingredients = self.ingredient_repo.get_all(db)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario Actual"
        
        # Activar cuadrícula visible
        ws.views.sheetView[0].showGridLines = True
        
        # Paleta de colores
        matcha_fill = PatternFill(start_color="7CB342", end_color="7CB342", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="336600")
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11)
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        double_bottom_border = Border(
            top=Side(style='thin', color='999999'),
            bottom=Side(style='double', color='000000')
        )

        # 1. Título del Reporte
        ws.merge_cells("A1:I1")
        ws["A1"] = "FitGenuss - Inventario de Ingredientes"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # Subtítulo (Fecha de exportación)
        ws.merge_cells("A2:I2")
        ws["A2"] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws["A2"].font = Font(name="Calibri", size=10, italic=True)
        ws.row_dimensions[2].height = 18

        # 2. Encabezados de Tabla
        headers = [
            "ID", "Nombre Ingrediente", "Categoría", "Stock Base", 
            "Unidad", "Precio Compra", "Costo Unitario", 
            "Stock Mínimo", "Valor Total"
        ]
        
        ws.append([]) # Fila vacía (Fila 3)
        ws.row_dimensions[4].height = 25
        
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.fill = matcha_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # 3. Rellenar Datos
        start_row = 5
        for ing in ingredients:
            total_val = ing.stock_actual * ing.costo_unitario
            row_data = [
                ing.id,
                ing.name,
                ing.category,
                ing.stock_actual,
                ing.unit_base,
                ing.precio_compra,
                ing.costo_unitario,
                ing.stock_minimo,
                total_val
            ]
            ws.append(row_data)
            current_row = ws.max_row
            ws.row_dimensions[current_row].height = 20
            
            # Aplicar estilos por celda
            for col_idx in range(1, 10):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                
                # Alineaciones y formatos
                if col_idx in [1]:  # ID
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx in [2, 3, 5]:  # Nombre, Categoria, Unidad
                    cell.alignment = Alignment(horizontal="left")
                elif col_idx in [4, 8]:  # Stocks
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0.00"
                elif col_idx in [6, 7, 9]:  # Valores monetarios
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = f'"${CURRENCY}"#,##0.00'

        end_row = ws.max_row
        
        # 4. Fila de Totales
        ws.append([]) # Fila intermedia
        total_row = ws.max_row + 1
        ws.row_dimensions[total_row].height = 22
        
        ws.cell(row=total_row, column=2, value="TOTAL VALORIZADO").font = bold_font
        ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="left")
        
        # Fórmula SUM para la valorización total
        val_cell = ws.cell(row=total_row, column=9, value=f"=SUM(I{start_row}:I{end_row})")
        val_cell.font = bold_font
        val_cell.border = double_bottom_border
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.number_format = f'"${CURRENCY}"#,##0.00'
        
        # Aplicar borde superior/inferior a toda la fila de totales
        for col_idx in range(1, 10):
            ws.cell(row=total_row, column=col_idx).border = double_bottom_border

        # Autoajustar anchos de columnas
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Evitar contar el largo de la fila de título mezclada
                if cell.row == 1:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Guardar en buffer binario
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_movements_excel(self, db: Session) -> bytes:
        """
        Genera el Kárdex de movimientos completo en formato Excel (.xlsx).
        """
        movements = self.movement_repo.get_all(db)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kárdex de Movimientos"
        ws.views.sheetView[0].showGridLines = True
        
        # Estilos
        peach_fill = PatternFill(start_color="FF8A80", end_color="FF8A80", fill_type="solid") # Rosa Pastel
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="990000")
        regular_font = Font(name="Calibri", size=11)
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        # Título
        ws.merge_cells("A1:G1")
        ws["A1"] = "FitGenuss - Historial de Movimientos de Inventario"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30
        
        ws.merge_cells("A2:G2")
        ws["A2"] = f"Exportado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws["A2"].font = Font(name="Calibri", size=10, italic=True)
        ws.row_dimensions[2].height = 18

        headers = ["ID Mov.", "Fecha y Hora", "Tipo de Movimiento", "Ingrediente", "Cant. Reg.", "Unidad", "Cant. Base Normalizada", "Observación"]
        ws.append([]) # Fila vacía
        ws.row_dimensions[4].height = 25
        
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.fill = peach_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for mov in movements:
            ing_name = mov.ingredient.name if mov.ingredient else f"Ingrediente ID {mov.ingredient_id}"
            ing_unit_base = mov.ingredient.unit_base if mov.ingredient else ""
            
            row_data = [
                mov.id,
                mov.date.strftime("%Y-%m-%d %H:%M:%S"),
                mov.type_movement.replace("_", " "),
                ing_name,
                mov.quantity,
                mov.unit,
                f"{mov.quantity_base:.2f} {ing_unit_base}",
                mov.observation or "-"
            ]
            ws.append(row_data)
            current_row = ws.max_row
            ws.row_dimensions[current_row].height = 20
            
            for col_idx in range(1, 9):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = regular_font
                cell.border = thin_border
                if col_idx in [1, 2, 6]:
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx in [3, 4, 8]:
                    cell.alignment = Alignment(horizontal="left")
                elif col_idx in [5, 7]:
                    cell.alignment = Alignment(horizontal="right")

        # Autoajustar columnas
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # --- REPORTES EN PDF (FPDF2) ---
    def generate_inventory_pdf(self, db: Session) -> bytes:
        """
        Genera un reporte ejecutivo en PDF del inventario actual.
        """
        ingredients = self.ingredient_repo.get_all(db)
        
        pdf = FitGenussPDF("REPORTE EJECUTIVO DE INVENTARIO VALORIZADO")
        pdf.alias_nb_pages()
        pdf.add_page()
        pdf.set_margins(10, 15, 10)
        
        # Encabezado de la tabla de datos
        pdf.set_y(35)
        pdf.set_fill_color(124, 179, 66) # Verde Matcha
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("helvetica", "B", 9)
        
        # Definición de anchos de columnas (Suma total 190)
        # ID: 10, Nombre: 48, Categoria: 32, Stock Base: 25, Unidad: 15, Costo Unit: 20, Minimo: 18, Total: 22
        cols = [10, 48, 32, 25, 15, 20, 18, 22]
        headers = ["ID", "Ingrediente", "Categoría", "Stock Base", "Unid.", "C. Unit.", "S. Min", "Val. Total"]
        
        for width, header in zip(cols, headers):
            pdf.cell(width, 7, header, border=1, align="C", fill=True)
        pdf.ln()

        # Rellenar filas
        pdf.set_text_color(50, 50, 50)
        pdf.set_font("helvetica", "", 8)
        
        alternating_color = False
        total_valuation = 0.0
        
        for ing in ingredients:
            val_total = ing.stock_actual * ing.costo_unitario
            total_valuation += val_total
            
            # Alternar color de fondo de filas
            if alternating_color:
                pdf.set_fill_color(245, 247, 245)
            else:
                pdf.set_fill_color(255, 255, 255)
            
            pdf.cell(cols[0], 6, str(ing.id), border=1, align="C", fill=True)
            pdf.cell(cols[1], 6, ing.name[:25], border=1, align="L", fill=True)
            pdf.cell(cols[2], 6, ing.category[:18], border=1, align="L", fill=True)
            pdf.cell(cols[3], 6, f"{ing.stock_actual:,.2f}", border=1, align="R", fill=True)
            pdf.cell(cols[4], 6, ing.unit_base, border=1, align="C", fill=True)
            pdf.cell(cols[5], 6, f"{CURRENCY} {ing.costo_unitario:,.3f}", border=1, align="R", fill=True)
            pdf.cell(cols[6], 6, f"{ing.stock_minimo:,.1f}", border=1, align="R", fill=True)
            pdf.cell(cols[7], 6, f"{CURRENCY} {val_total:,.2f}", border=1, align="R", fill=True)
            pdf.ln()
            
            alternating_color = not alternating_color

        # Fila de Totales
        pdf.set_font("helvetica", "B", 9)
        pdf.set_fill_color(230, 240, 220)
        # Combinar columnas anteriores a Valor Total
        ancho_previo = sum(cols[:-1])
        pdf.cell(ancho_previo, 7, "VALORIZACIÓN TOTAL DEL INVENTARIO", border=1, align="R", fill=True)
        pdf.cell(cols[-1], 7, f"{CURRENCY} {total_valuation:,.2f}", border=1, align="R", fill=True)

        return bytes(pdf.output())

    def generate_movements_pdf(self, db: Session) -> bytes:
        """
        Genera un reporte del historial de movimientos en formato PDF.
        """
        movements = self.movement_repo.get_all(db)
        
        pdf = FitGenussPDF("HISTORIAL DE MOVIMIENTOS - KÁRDEX COMPLETO")
        pdf.alias_nb_pages()
        pdf.add_page()
        pdf.set_margins(10, 15, 10)
        
        pdf.set_y(35)
        pdf.set_fill_color(255, 138, 128) # Rosa Pastel
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("helvetica", "B", 9)
        
        # Anchos de columnas (Suma 190)
        # ID: 12, Fecha: 30, Tipo Mov: 38, Ingrediente: 42, Cant. Reg: 18, Unid: 15, Cant. Base: 35
        cols = [12, 30, 38, 42, 18, 15, 35]
        headers = ["ID Mov.", "Fecha/Hora", "Tipo Movimiento", "Ingrediente", "Cant.", "Unid.", "Equivalencia Base"]
        
        for width, header in zip(cols, headers):
            pdf.cell(width, 7, header, border=1, align="C", fill=True)
        pdf.ln()

        # Datos
        pdf.set_text_color(50, 50, 50)
        pdf.set_font("helvetica", "", 8)
        
        alternating_color = False
        
        for mov in movements:
            ing_name = mov.ingredient.name if mov.ingredient else f"Ing. ID {mov.ingredient_id}"
            ing_unit_base = mov.ingredient.unit_base if mov.ingredient else ""
            
            if alternating_color:
                pdf.set_fill_color(253, 245, 245)
            else:
                pdf.set_fill_color(255, 255, 255)
                
            pdf.cell(cols[0], 6, str(mov.id), border=1, align="C", fill=True)
            pdf.cell(cols[1], 6, mov.date.strftime("%Y-%m-%d %H:%M"), border=1, align="C", fill=True)
            pdf.cell(cols[2], 6, mov.type_movement.replace("_", " "), border=1, align="L", fill=True)
            pdf.cell(cols[3], 6, ing_name[:22], border=1, align="L", fill=True)
            pdf.cell(cols[4], 6, f"{mov.quantity:.2f}", border=1, align="R", fill=True)
            pdf.cell(cols[5], 6, mov.unit, border=1, align="C", fill=True)
            pdf.cell(cols[6], 6, f"{mov.quantity_base:.1f} {ing_unit_base}", border=1, align="R", fill=True)
            pdf.ln()
            
            alternating_color = not alternating_color

        return bytes(pdf.output())
